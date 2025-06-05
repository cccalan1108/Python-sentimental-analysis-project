from selenium import webdriver
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from datetime import datetime, timezone
import time
import re
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import os
from openpyxl import load_workbook, Workbook
def save_post_to_excel(post_data, account):
    file_name = "threads.xlsx"

    if not os.path.exists(file_name):
        # 如果 Excel 不存在，建立一個新的
        wb = Workbook()
        ws = wb.active
        ws.title = account
        ws.append(list(post_data.keys()))  # 寫入欄位名稱
        ws.append(list(post_data.values()))
        wb.save(file_name)
        print(f"✅ 建立新檔案並新增一筆資料到 {file_name}")
    else:
        # 如果存在，讀取後追加
        wb = load_workbook(file_name)
        if account in wb.sheetnames:
            ws = wb[account]
        else:
            ws = wb.create_sheet(title=account)
            ws.append(list(post_data.keys()))  # 新 sheet 要寫欄位名稱

        ws.append(list(post_data.values()))
        wb.save(file_name)
        print(f"✅ 已新增一筆資料到 {file_name} 的 {account} 工作表")

def scrape_threads(account, password, target_post_count):
    results = []
    collected_posts = 0
    scroll_count = 0
    max_scrolls = 30

    options = Options()
    driver = webdriver.Chrome(options=options)

    try:
        driver.get("https://www.threads.net/login")
        time.sleep(5)

        username_input = driver.find_element(By.CSS_SELECTOR, 'input[autocomplete="username"]')
        username_input.send_keys(account)

        password_input = driver.find_element(By.CSS_SELECTOR, 'input[autocomplete="current-password"]')
        password_input.send_keys(password)

        login_button = driver.find_element(By.XPATH, '//div[text()="登入"]/ancestor::div[@role="button"]')
        login_button.click()
        time.sleep(20)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "time")))


        while collected_posts < target_post_count and scroll_count < max_scrolls:
            current_post_blocks = driver.find_elements(By.XPATH, '//div[@data-pressable-container="true"]')

            if len(current_post_blocks) > collected_posts:
                for idx, post in enumerate(current_post_blocks[collected_posts:], collected_posts + 1):
                    soup = BeautifulSoup(post.get_attribute("innerHTML"), "html.parser")

                    author_tag = soup.find("a", href=lambda x: x and x.startswith("/@"))
                    author = author_tag.get_text(strip=True) if author_tag else "未知"

                    time_tag = soup.find("time")
                    if time_tag and time_tag.has_attr("title"):
                        try:
                            time_title = time_tag["title"]
                            if "年" in time_title:
                                year = int(re.search(r'(\\d+)年', time_title).group(1))
                                month = int(re.search(r'(\\d+)月', time_title).group(1))
                                day = int(re.search(r'(\\d+)日', time_title).group(1))
                                time_part = re.search(r'(上午|下午)(\\d+):(\\d+)', time_title)
                                if time_part:
                                    am_pm = time_part.group(1)
                                    hour = int(time_part.group(2))
                                    minute = int(time_part.group(3))
                                    if am_pm == "下午" and hour < 12:
                                        hour += 12
                                    elif am_pm == "上午" and hour == 12:
                                        hour = 0
                                    post_time = f"{year}年{month:02d}月{day:02d}日 {hour:02d}:{minute:02d}"
                                else:
                                    post_time = "未知"
                            else:
                                parsed_date = datetime.strptime(time_title, "%A, %B %d, %Y at %I:%M %p")
                                post_time = parsed_date.strftime("%Y年%m月%d日 %H:%M")
                        except:
                            post_time = time_tag["datetime"] if time_tag.has_attr("datetime") else "未知"
                    else:
                        post_time = "未知"

                    photo, video = check_media_types(soup)
                    content_spans = soup.select('span[dir=\"auto\"]')
                    content = " ".join([s.get_text(strip=True) for s in content_spans if s.get_text(strip=True)])
                    result = extract_post_info(content)

                    detail_a_tag = soup.find("a", href=lambda x: x and "/post/" in x)
                    if detail_a_tag:
                        post_url = f"https://www.threads.net{detail_a_tag['href']}"
                        stats = open_post_detail(driver, post_url)

                        single_post = {
                            "author": author,
                            "post_time": post_time,
                            "topic": result["topic"],
                            "time_info": result["time"],
                            "content": result["content"],
                            "has_photo": photo,
                            "has_video": video,
                            "like_count": stats["like_count"],
                            "reply_count": stats["reply_count"],
                            "repost_count": stats["repost_count"],
                            "share_count": stats["share_count"],
                            "view_count": stats["view_count"],
                            "followers_count": stats["followers_count"],
                            "post_url": post_url,
                            "scrape_time": datetime.now().isoformat()
                        }

                        # 新增結果
                        results.append(single_post)
                        # 立刻寫入 Excel
                        save_post_to_excel(single_post, account)
                    else:
                        print("❌ 找不到詳細貼文連結，略過")

                    time.sleep(3)

            collected_posts = len(current_post_blocks)
            if collected_posts >= target_post_count:
                break

            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            print("向下滾動頁面...")
            time.sleep(5)

            try:
                previous_count = collected_posts
                WebDriverWait(driver, 10).until(
                    lambda d: len(d.find_elements(By.XPATH, '//div[@data-pressable-container=\"true\"]')) > previous_count
                )
            except:
                print("滾動後沒有加載新貼文")
            scroll_count += 1

        print(f"🎯 完成，總共收集 {collected_posts} 筆貼文")

        return results

    finally:
        driver.quit()

def extract_post_info(raw_text: str):
    """從貼文內容中提取資訊"""
    # 刪除「翻譯」與結尾數字
    raw_text = re.sub(r"翻譯.*?$", "", raw_text)
    raw_text = re.sub(r"\d+\s+\d*\s*\d*\s*\d*\s*$", "", raw_text).strip()

    parts = raw_text.split()

    if len(parts) < 3:
        return {"author": "", "topic": "", "time": "", "content": ""}

    # 找出時間欄位（以「分鐘」、「小時」、「天」等單位判斷）
    time_index = next((i for i, word in enumerate(parts) if re.search(r'(分鐘|小時|天)$', word)), None)

    if time_index is None or time_index < 1:
        return {"author": "", "topic": "", "time": "", "content": ""}

    author = parts[0]
    
    # 主題是從作者後到時間前的所有內容
    topic = " ".join(parts[1:time_index])
    
    time_str = parts[time_index]
    
    # 內容是時間欄位之後的所有文字
    content = " ".join(parts[time_index + 1:])

    return {
        "author": author,
        "topic": topic,
        "time": time_str,
        "content": content
    }

def check_media_types(soup):
    """檢查貼文是否包含圖片或影片"""
    photo = "N"
    video = "N"

    # 檢查圖片（排除大頭貼）
    img_tags = soup.find_all("img")
    for img in img_tags:
        alt = img.get("alt", "")
        if "大頭貼照" not in alt:
            photo = "Y"
            break

    # 檢查影片
    if soup.find("video"):
        video = "Y"

    return photo, video

def open_post_detail(driver, url: str):
    """開啟貼文詳情頁面並獲取統計資訊"""
    # 開新分頁
    driver.execute_script("window.open(arguments[0], '_blank');", url)
    driver.switch_to.window(driver.window_handles[1])
    time.sleep(5)

    # BeautifulSoup 處理頁面
    soup = BeautifulSoup(driver.page_source, "html.parser")

    # 擷取數據統計數字
    def get_stat_by_index(index):
        spans = soup.find_all("span", class_=re.compile("x17qophe"))
        try:
            return spans[index].get_text(strip=True)
        except IndexError:
            return "0"

    like_count = get_stat_by_index(0)
    reply_count = get_stat_by_index(1)
    repost_count = get_stat_by_index(2)
    share_count = get_stat_by_index(3)

    print(f"❤️ 按讚數：{like_count}")
    print(f"💬 留言數：{reply_count}")
    print(f"🔁 轉發數：{repost_count}")
    print(f"📤 分享數：{share_count}")

    # 爬取「瀏覽次數」- 根據新提供的HTML結構
    try:
        # 尋找包含「次瀏覽」文字的span
        view_count_span = soup.find("span", string=lambda s: s and ("次瀏覽" in s))
        
        if view_count_span:
            # 提取瀏覽次數（移除「次瀏覽」文字）
            view_count_text = view_count_span.get_text(strip=True)
            view_count = view_count_text.replace("次瀏覽", "")
        else:
            # 備用方法：尋找特定結構
            views_div = soup.find("div", class_=lambda c: c and "x6s0dn4" in c and "xfex06f" in c)
            if views_div:
                view_span = views_div.find("span", string=lambda s: s and ("次瀏覽" in s or "views" in s.lower()))
                if view_span:
                    view_count_text = view_span.get_text(strip=True)
                    view_count = view_count_text.replace("次瀏覽", "").replace("views", "").strip()
                else:
                    view_count = "未知"
            else:
                view_count = "未知"
        
        print("👁️ 觀看次數：", view_count)
    except Exception as e:
        print("⚠️ 觀看次數解析失敗：", e)
        view_count = "0"
        
    try:
        time_tag = soup.find("time")
        author_href, author_name = "未知", "未知"
        author_username = "未知"

        if time_tag:
            # 找到 time 所在區塊的父層（例如含 post metadata 的區塊）
            block = time_tag.find_parent("div")
            if block:
                # 在這區塊內尋找 <a href="/@xxx">
                author_link = block.find("a", href=re.compile(r"^/@"), attrs={"role": "link"})
                if author_link:
                    author_href = author_link["href"]
                    author_username = author_href.split("/")[1].lstrip("@")
        print(f"✍️ 發文者帳號：{author_username}")
        print(f"🔗 個人連結：https://www.threads.net/@{author_username}")
    except Exception as e:
        print("⚠️ 發文者解析失敗：", e)
        
    followers_count = "0"
    try:
        # 在新分頁中打開發文者的主頁
        driver.execute_script("window.open(arguments[0], '_blank');", f"https://www.threads.net/@{author_username}")
        
        # 切換到發文者主頁分頁（window_handles[2]，因為當前是[1]）
        driver.switch_to.window(driver.window_handles[2])
        
        # 等待頁面加載
        time.sleep(5)
        
        # 抓取追蹤者數量
        followers_count = get_followers_count(driver)
        print(f"👥 追蹤者數量：{followers_count}")
        
        # 關閉發文者主頁分頁
        driver.close()
        
        # 回到貼文分頁
        driver.switch_to.window(driver.window_handles[1])
        
    except Exception as e:
        print(f"⚠️ 抓取追蹤者數量時出錯：{e}")
    
    # 關閉貼文頁並回到主分頁
    driver.close()
    driver.switch_to.window(driver.window_handles[0])
    time.sleep(1)
    
    return {
        "like_count": like_count,
        "reply_count": reply_count,
        "repost_count": repost_count,
        "share_count": share_count,
        "view_count": view_count,
        "followers_count": followers_count
    }

def get_followers_count(driver, soup=None):
    """
    從目前頁面抓取追蹤者數量（支援中英文）

    :param driver: WebDriver 實例
    :param soup: 可選，已解析的 BeautifulSoup 物件。如果沒提供會從 driver.page_source 創建
    :return: 追蹤者數量字串
    """
    try:
        if soup is None:
            soup = BeautifulSoup(driver.page_source, 'html.parser')

        span_tags = soup.find_all("span")

        for span in span_tags:
            text = span.get_text(strip=True)
            # 支援 "followers" 或 "位粉絲"
            if "followers" in text.lower() or "位粉絲" in text:
                inner_span = span.find("span")
                if inner_span and inner_span.has_attr("title"):
                    return inner_span["title"].replace(",", "")
                else:
                    match = re.search(r"(\d+(?:,\d+)*)", text)
                    if match:
                        return match.group(1).replace(",", "")
        
        return "0"

    except Exception as e:
        print(f"⚠️ 抓取追蹤者數量時出錯：{e}")
        return "0"

accounts = []

if __name__ == "__main__":
        account = ""
        target_post_count = 400
        
        scrape_threads(account, "", target_post_count)
